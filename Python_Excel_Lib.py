import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import re
from uuid import uuid4

class Excel_WorkBook:
    def __init__(self, str_path_file_excel, str_name_sheet="Sheet"):
        """
        Initialize Excel workbook with enhanced error handling
        """
        self.str_path_file_excel = str_path_file_excel
        try:
            if os.path.exists(str_path_file_excel):
                self.workbook = openpyxl.load_workbook(str_path_file_excel)
                if self.__check_name_sheet__(str_name_sheet):
                    print(f"Sheet '{str_name_sheet}' already exists.")
                else:
                    self.workbook.create_sheet(title=str_name_sheet)
                    print(f"Created new sheet '{str_name_sheet}'.")
                print("Excel file opened successfully.")
            else:
                self.workbook = openpyxl.Workbook()
                # Remove default sheet and create new one with specified name
                default_sheet = self.workbook.active
                self.workbook.remove(default_sheet)
                self.workbook.create_sheet(title=str_name_sheet)
                print(f"Created new Excel file with sheet '{str_name_sheet}'.")
            self.active_sheet = str_name_sheet
        except Exception as e:
            print(f"Error initializing workbook: {str(e)}")
            raise

    def __check_name_sheet__(self, str_name_sheet):
        """Check if sheet name exists"""
        return str_name_sheet in self.workbook.sheetnames

    def __validate_cell_reference(self, cell_ref):
        """Validate cell reference format (e.g., A1, B2)"""
        pattern = r'^[A-Z]+[1-9][0-9]*$'
        return bool(re.match(pattern, cell_ref.upper()))

    def create_sheet(self, str_name_sheet, overwrite=False):
        """Create new sheet with option to overwrite"""
        try:
            if not overwrite and self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' already exists.")
                return False
            if overwrite and self.__check_name_sheet__(str_name_sheet):
                self.workbook.remove(self.workbook[str_name_sheet])
            self.workbook.create_sheet(title=str_name_sheet)
            print(f"Created sheet '{str_name_sheet}'.")
            return True
        except Exception as e:
            print(f"Error creating sheet: {str(e)}")
            return False

    def set_active_sheet(self, str_name_sheet):
        """Set the active sheet"""
        if self.__check_name_sheet__(str_name_sheet):
            self.active_sheet = str_name_sheet
            self.workbook.active = self.workbook[str_name_sheet]
            return True
        print(f"Sheet '{str_name_sheet}' does not exist.")
        return False

    def get_sheet(self, str_name_sheet):
        """Return specified sheet object"""
        if self.__check_name_sheet__(str_name_sheet):
            return self.workbook[str_name_sheet]
        print(f"Sheet '{str_name_sheet}' does not exist.")
        return None

    def get_sheet_names(self):
        """Return list of sheet names"""
        return self.workbook.sheetnames

    def write_column(self, str_name_sheet, column, list_content, start_row=1):
        """Write data to a column (accepts both letter and number column index)"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            # Convert column letter to number if needed
            if isinstance(column, str):
                column = openpyxl.utils.column_index_from_string(column)
            
            for index, data in enumerate(list_content, start=start_row):
                sheet.cell(row=index, column=column).value = data
            return True
        except Exception as e:
            print(f"Error writing column: {str(e)}")
            return False

    def write_row(self, str_name_sheet, row, list_content, start_column=1):
        """Write data to a row"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            for index, data in enumerate(list_content, start=start_column):
                sheet.cell(row=row, column=index).value = data
            return True
        except Exception as e:
            print(f"Error writing row: {str(e)}")
            return False

    def write_cell(self, str_name_sheet, cell_ref, content):
        """Write to a specific cell (accepts both A1-style and row/column coordinates)"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            if isinstance(cell_ref, str) and self.__validate_cell_reference(cell_ref):
                sheet[cell_ref].value = content
            elif isinstance(cell_ref, tuple) and len(cell_ref) == 2:
                sheet.cell(row=cell_ref[0], column=cell_ref[1]).value = content
            else:
                print("Invalid cell reference format.")
                return False
            return True
        except Exception as e:
            print(f"Error writing cell: {str(e)}")
            return False

    def read_cell(self, str_name_sheet, cell_ref):
        """Read from a specific cell"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return None
            
            sheet = self.workbook[str_name_sheet]
            if isinstance(cell_ref, str) and self.__validate_cell_reference(cell_ref):
                return sheet[cell_ref].value
            elif isinstance(cell_ref, tuple) and len(cell_ref) == 2:
                return sheet.cell(row=cell_ref[0], column=cell_ref[1]).value
            print("Invalid cell reference format.")
            return None
        except Exception as e:
            print(f"Error reading cell: {str(e)}")
            return None

    def read_range(self, str_name_sheet, start_cell, end_cell):
        """Read a range of cells (e.g., 'A1:C3')"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return None
            
            sheet = self.workbook[str_name_sheet]
            return [[cell.value for cell in row] for row in sheet[start_cell:end_cell]]
        except Exception as e:
            print(f"Error reading range: {str(e)}")
            return None

    def set_column_width(self, str_name_sheet, column, width):
        """Set column width"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            if isinstance(column, int):
                column = get_column_letter(column)
            sheet.column_dimensions[column].width = width
            return True
        except Exception as e:
            print(f"Error setting column width: {str(e)}")
            return False

    def set_row_height(self, str_name_sheet, row, height):
        """Set row height"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            sheet.row_dimensions[row].height = height
            return True
        except Exception as e:
            print(f"Error setting row height: {str(e)}")
            return False

    def format_cells(self, str_name_sheet, cell_range, pattern_fill=None, font=None, border=None, alignment=None, number_format=None):
        """Format cells with multiple style options"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            for row in sheet[cell_range]:
                for cell in row:
                    if pattern_fill:
                        cell.fill = pattern_fill
                    if font:
                        cell.font = font
                    if border:
                        cell.border = border
                    if alignment:
                        cell.alignment = alignment
                    if number_format:
                        cell.number_format = number_format
            return True
        except Exception as e:
            print(f"Error formatting cells: {str(e)}")
            return False

    def insert_image(self, str_name_sheet, cell_ref, image_path, scale_width=1.0, scale_height=1.0):
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' not exists.")
                return False
            if not os.path.exists(image_path):
                print(f"File ảnh '{image_path}' not exists.")
                return False
                
            sheet = self.workbook[str_name_sheet]
            img = Image(image_path)
            # Điều chỉnh kích thước ảnh
            img.width = img.width * scale_width
            img.height = img.height * scale_height
            sheet.add_image(img, cell_ref)
            return True
        except Exception as e:
            print(f"Error inserting image: {str(e)}")
            return False

    def set_formula(self, str_name_sheet, cell_ref, formula):
        """Set Excel formula in a cell"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            if isinstance(cell_ref, str) and self.__validate_cell_reference(cell_ref):
                sheet[cell_ref].value = formula
            elif isinstance(cell_ref, tuple) and len(cell_ref) == 2:
                sheet.cell(row=cell_ref[0], column=cell_ref[1]).value = formula
            else:
                print("Invalid cell reference format.")
                return False
            return True
        except Exception as e:
            print(f"Error setting formula: {str(e)}")
            return False

    def freeze_panes(self, str_name_sheet, cell_ref):
        """Freeze panes at specified cell"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            sheet.freeze_panes = cell_ref
            return True
        except Exception as e:
            print(f"Error setting freeze panes: {str(e)}")
            return False

    def add_sort_filter(self, str_name_sheet, cell_range):
        """Add sort and filter to specified range"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            sheet.auto_filter.ref = cell_range
            return True
        except Exception as e:
            print(f"Error adding sort filter: {str(e)}")
            return False

    def merge_cells(self, str_name_sheet, cell_range):
        """Merge cells in specified range"""
        try:
            if not self.__check_name_sheet__(str_name_sheet):
                print(f"Sheet '{str_name_sheet}' does not exist.")
                return False
            
            sheet = self.workbook[str_name_sheet]
            sheet.merge_cells(cell_range)
            return True
        except Exception as e:
            print(f"Error merging cells: {str(e)}")
            return False

    def save(self, path_save=None):
        """Save workbook to specified path or original path"""
        try:
            path = path_save if path_save else self.str_path_file_excel
            self.workbook.save(path)
            print(f"Workbook saved to {path}")
            return True
        except Exception as e:
            print(f"Error saving workbook: {str(e)}")
            return False

    def close(self):
        """Close workbook"""
        try:
            self.workbook.close()
            print("Workbook closed successfully.")
            return True
        except Exception as e:
            print(f"Error closing workbook: {str(e)}")
            return False

    def check_last_data_cell(self, str_name_sheet):
            """
            Trong thư viện openpyxl, xác định ô cuối cùng có dữ liệu bằng cách sử dụng thuộc tính max_row và max_column của sheet, 
            kết hợp với việc kiểm tra thực tế các ô từ cuối sheet trở lên. Tuy nhiên, max_row và max_column chỉ cho biết phạm vi lớn nhất mà sheet đã sử dụng, 
            chứ không đảm bảo rằng mọi ô trong phạm vi đó đều có dữ liệu. 
            Vì vậy, chúng ta cần một phương pháp quét ngược từ cuối sheet để tìm chính xác ô cuối cùng có dữ liệu, và sau đó kiểm tra các ô tiếp theo một cách hiệu quả hơn.
            """
            try:
                if not self.__check_name_sheet__(str_name_sheet):
                    print(f"Sheet '{str_name_sheet}' does not exist.")
                    return None
                
                sheet = self.workbook[str_name_sheet]
                max_row = sheet.max_row or 1
                max_col = sheet.max_column or 1

                # Start from the bottom-right cell and scan backwards to find the last cell with data
                last_row = 0
                last_col = 0
                for row in range(max_row, 0, -1):
                    for col in range(max_col, 0, -1):
                        if sheet.cell(row=row, column=col).value is not None:
                            last_row = row
                            last_col = col
                            break
                    if last_row > 0:
                        break

                if last_row == 0 or last_col == 0:
                    print("No data found in the sheet.")
                    return (0, 0, True)

                # Check a reasonable range after the last cell (e.g., next 1000 rows or columns)
                check_range = 1000
                is_clean = True
                # Check rows after last_row
                for row in range(last_row + 1, min(last_row + check_range + 1, max_row + 1)):
                    for col in range(1, max_col + 1):
                        if sheet.cell(row=row, column=col).value is not None:
                            is_clean = False
                            print(f"Unexpected data found at row {row}, column {col}")
                            break
                    if not is_clean:
                        break
                # Check columns after last_col
                for col in range(last_col + 1, min(last_col + check_range + 1, max_col + 1)):
                    for row in range(1, last_row + 1):
                        if sheet.cell(row=row, column=col).value is not None:
                            is_clean = False
                            print(f"Unexpected data found at row {row}, column {col}")
                            break
                    if not is_clean:
                        break

                return (last_row, last_col, is_clean)
            except Exception as e:
                print(f"Error checking last data cell: {str(e)}")
                return None