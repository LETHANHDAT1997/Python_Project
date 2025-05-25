import openpyxl
from openpyxl.styles import Font, PatternFill, GradientFill, Border, Side, Alignment, Protection, NamedStyle, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, IconSetRule, DataBarRule
from uuid import uuid4

class Excel_Style:
    def __init__(self):
        """
        Initialize Excel_Style with a dictionary to store named styles
        """
        self.named_styles = {}
        print("Excel_Style initialized successfully.")

    def create_font(self, name='Calibri', size=11, bold=False, italic=False, underline=None, color='FF000000', 
                    strikethrough=False, vert_align=None):
        """
        Create a Font object with specified properties
        """
        try:
            font = Font(
                name=name,
                size=size,
                bold=bold,
                italic=italic,
                underline=underline,  # e.g., 'single', 'double'
                color=color,         # RGB or theme color
                strikethrough=strikethrough,
                vertAlign=vert_align  # e.g., 'superscript', 'subscript'
            )
            return font
        except Exception as e:
            print(f"Error creating font: {str(e)}")
            return None

    def create_pattern_fill(self, fill_type='solid', start_color='FFFFFF', end_color='FFFFFF'):
        """
        Create a PatternFill object
        """
        try:
            fill = PatternFill(
                fill_type=fill_type,  # e.g., 'solid', 'darkDown', 'lightGrid'
                start_color=start_color,
                end_color=end_color
            )
            return fill
        except Exception as e:
            print(f"Error creating pattern fill: {str(e)}")
            return None

    def create_gradient_fill(self, fill_type='linear', stop=('FFFFFF', '000000'), degree=0):
        """
        Create a GradientFill object
        """
        try:
            fill = GradientFill(
                type=fill_type,  # 'linear' or 'path'
                stop=stop,       # List of colors
                degree=degree    # Angle for linear gradient
            )
            return fill
        except Exception as e:
            print(f"Error creating gradient fill: {str(e)}")
            return None

    def create_border(self, left_style=None, right_style=None, top_style=None, bottom_style=None, 
                     left_color='FF000000', right_color='FF000000', top_color='FF000000', bottom_color='FF000000', 
                     diagonal=False, diagonal_style=None, diagonal_color='FF000000'):
        """
        Create a Border object
        """
        try:
            border = Border(
                left=Side(border_style=left_style, color=left_color) if left_style else None,
                right=Side(border_style=right_style, color=right_color) if right_style else None,
                top=Side(border_style=top_style, color=top_color) if top_style else None,
                bottom=Side(border_style=bottom_style, color=bottom_color) if bottom_style else None,
                diagonal=Side(border_style=diagonal_style, color=diagonal_color) if diagonal else None,
                diagonalUp=diagonal,
                diagonalDown=diagonal
            )
            return border
        except Exception as e:
            print(f"Error creating border: {str(e)}")
            return None

    def create_alignment(self, horizontal='general', vertical='bottom', text_rotation=0, wrap_text=False, 
                        shrink_to_fit=False, indent=0):
        """
        Create an Alignment object
        """
        try:
            alignment = Alignment(
                horizontal=horizontal,  # e.g., 'center', 'left', 'right'
                vertical=vertical,      # e.g., 'center', 'top', 'bottom'
                text_rotation=text_rotation,
                wrap_text=wrap_text,
                shrink_to_fit=shrink_to_fit,
                indent=indent
            )
            return alignment
        except Exception as e:
            print(f"Error creating alignment: {str(e)}")
            return None

    def create_protection(self, locked=True, hidden=False):
        """
        Create a Protection object
        """
        try:
            protection = Protection(
                locked=locked,
                hidden=hidden
            )
            return protection
        except Exception as e:
            print(f"Error creating protection: {str(e)}")
            return None

    def create_named_style(self, name, font=None, fill=None, border=None, alignment=None, number_format=None, 
                          protection=None):
        """
        Create and store a NamedStyle object
        """
        try:
            style = NamedStyle(name=name)
            if font:
                style.font = font
            if fill:
                style.fill = fill
            if border:
                style.border = border
            if alignment:
                style.alignment = alignment
            if number_format:
                style.number_format = number_format
            if protection:
                style.protection = protection
            self.named_styles[name] = style
            return style
        except Exception as e:
            print(f"Error creating named style: {str(e)}")
            return None

    def apply_named_style(self, workbook, style_name):
        """
        Add a named style to the workbook
        """
        try:
            if style_name not in self.named_styles:
                print(f"Named style '{style_name}' does not exist.")
                return False
            workbook.add_named_style(self.named_styles[style_name])
            return True
        except Exception as e:
            print(f"Error applying named style: {str(e)}")
            return False

    def create_conditional_formatting(self, condition_type, cell_range, workbook, sheet_name, 
                                     operator=None, value=None, font=None, fill=None, 
                                     color_scale_colors=None, icon_set=None, data_bar_color=None):
        """
        Create and apply conditional formatting
        """
        try:
            sheet = workbook[sheet_name]
            if condition_type == 'cellIs':
                dxf = DifferentialStyle(font=font, fill=fill)
                rule = Rule(type=condition_type, operator=operator, formula=[value], dxf=dxf)
            elif condition_type == 'colorScale':
                rule = ColorScaleRule(
                    start_type='min', start_color=color_scale_colors[0],
                    mid_type='percentile', mid_value=50, mid_color=color_scale_colors[1],
                    end_type='max', end_color=color_scale_colors[2]
                )
            elif condition_type == 'iconSet':
                rule = IconSetRule(icon_style=icon_set, type='3Arrows', showValue=False)
            elif condition_type == 'dataBar':
                rule = DataBarRule(start_type='min', end_type='max', color=data_bar_color)
            else:
                print(f"Unsupported conditional formatting type: {condition_type}")
                return False
            
            sheet.conditional_formatting.add(cell_range, rule)
            return True
        except Exception as e:
            print(f"Error creating conditional formatting: {str(e)}")
            return False

    def get_color(self, color_type='rgb', value='FF000000'):
        """
        Return a color object (RGB, theme, or indexed)
        """
        try:
            if color_type == 'rgb':
                return colors.Color(rgb=value)
            elif color_type == 'theme':
                return getattr(colors, value.upper(), colors.BLACK)
            elif color_type == 'indexed':
                return colors.Color(indexed=int(value))
            print(f"Unsupported color type: {color_type}")
            return None
        except Exception as e:
            print(f"Error creating color: {str(e)}")
            return None
        
    def get_cell_style(self, workbook, sheet_name, cell_ref):
        """
        Retrieve style information (font, fill, border, alignment, number_format, protection) for a specific cell
        """
        try:
            if sheet_name not in workbook.sheetnames:
                print(f"Sheet '{sheet_name}' does not exist.")
                return None
            
            sheet = workbook[sheet_name]
            cell = sheet[cell_ref] if isinstance(cell_ref, str) else sheet.cell(row=cell_ref[0], column=cell_ref[1])
            
            style_info = {
                'font': {
                    'name': cell.font.name if cell.font else None,
                    'size': cell.font.size if cell.font else None,
                    'bold': cell.font.bold if cell.font else None,
                    'italic': cell.font.italic if cell.font else None,
                    'underline': cell.font.underline if cell.font else None,
                    'color': cell.font.color.rgb if cell.font and cell.font.color else None,
                    'strikethrough': cell.font.strikethrough if cell.font else None,
                    'vertAlign': cell.font.vertAlign if cell.font else None
                },
                'fill': {
                    'fill_type': cell.fill.fill_type if cell.fill else None,
                    'start_color': cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None,
                    'end_color': cell.fill.end_color.rgb if cell.fill and cell.fill.end_color else None
                },
                'border': {
                    'left_style': cell.border.left.border_style if cell.border and cell.border.left else None,
                    'left_color': cell.border.left.color.rgb if cell.border and cell.border.left and cell.border.left.color else None,
                    'right_style': cell.border.right.border_style if cell.border and cell.border.right else None,
                    'right_color': cell.border.right.color.rgb if cell.border and cell.border.right and cell.border.right.color else None,
                    'top_style': cell.border.top.border_style if cell.border and cell.border.top else None,
                    'top_color': cell.border.top.color.rgb if cell.border and cell.border.top and cell.border.top.color else None,
                    'bottom_style': cell.border.bottom.border_style if cell.border and cell.border.bottom else None,
                    'bottom_color': cell.border.bottom.color.rgb if cell.border and cell.border.bottom and cell.border.bottom.color else None,
                    'diagonal': cell.border.diagonalUp or cell.border.diagonalDown if cell.border else None,
                    'diagonal_style': cell.border.diagonal.border_style if cell.border and cell.border.diagonal else None,
                    'diagonal_color': cell.border.diagonal.color.rgb if cell.border and cell.border.diagonal and cell.border.diagonal.color else None
                },
                'alignment': {
                    'horizontal': cell.alignment.horizontal if cell.alignment else None,
                    'vertical': cell.alignment.vertical if cell.alignment else None,
                    'text_rotation': cell.alignment.text_rotation if cell.alignment else None,
                    'wrap_text': cell.alignment.wrap_text if cell.alignment else None,
                    'shrink_to_fit': cell.alignment.shrink_to_fit if cell.alignment else None,
                    'indent': cell.alignment.indent if cell.alignment else None
                },
                'number_format': cell.number_format if cell.number_format else None,
                'protection': {
                    'locked': cell.protection.locked if cell.protection else None,
                    'hidden': cell.protection.hidden if cell.protection else None
                }
            }
            return style_info
        except Exception as e:
            print(f"Error retrieving cell style: {str(e)}")
            return None

if __name__ == "__main__":
    # Example usage
    style_manager = Excel_Style()

    # Create sample styles
    font = style_manager.create_font(name='Tahoma', size=12, bold=True, color='FF0000')
    pattern_fill = style_manager.create_pattern_fill(fill_type='solid', start_color='FFFF00')
    gradient_fill = style_manager.create_gradient_fill(fill_type='linear', stop=('FFFFFF', '0000FF'), degree=45)
    border = style_manager.create_border(left_style='thin', right_style='thin', top_style='thin', bottom_style='thin')
    alignment = style_manager.create_alignment(horizontal='center', vertical='center', wrap_text=True)
    protection = style_manager.create_protection(locked=True, hidden=False)

    # Create and store a named style
    header_style = style_manager.create_named_style(
        name='header_style',
        font=font,
        fill=pattern_fill,
        border=border,
        alignment=alignment,
        number_format='#,##0.00',
        protection=protection
    )

    # Example with a workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'TestSheet'

    # Apply named style to workbook
    style_manager.apply_named_style(workbook, 'header_style')
    sheet['A1'].style = 'header_style'
    sheet['A1'].value = 'Header'

    # Apply conditional formatting
    style_manager.create_conditional_formatting(
        condition_type='cellIs',
        cell_range='B1:B10',
        workbook=workbook,
        sheet_name='TestSheet',
        operator='greaterThan',
        value='100',
        font=Font(color='FF0000'),
        fill=PatternFill(start_color='FFFF00', fill_type='solid')
    )

    # Retrieve and print cell style for A1
    style_info = style_manager.get_cell_style(workbook, 'TestSheet', 'A1')
    if style_info:
        print(f"Cell A1 Style:\n{style_info['font']}\n{style_info['fill']}\n{style_info['border']}\n{style_info['alignment']}\n{style_info['number_format']}\n{style_info['protection']}")

    # Save workbook
    workbook.save('style_test.xlsx')
    print("Style test workbook saved.")